import cv2
import math
import time
import numpy as np
import os
from openpyxl import Workbook, load_workbook

# Speed limit in km/hr
limit = 80  

# Folder to store records and images
traffic_record_folder_name = "TrafficRecord"

# Create necessary folders
if not os.path.exists(traffic_record_folder_name):
    os.makedirs(traffic_record_folder_name)
    os.makedirs(os.path.join(traffic_record_folder_name, "exceeded"))

# Speed record file
speed_record_file_location = os.path.join(traffic_record_folder_name, "SpeedRecord.xlsx")

# Initialize Excel file with openpyxl
def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(speed_record_file_location):
        wb = Workbook()
        ws = wb.active
        ws.title = "Speed Records"
        ws.append(["ID", "Speed (km/h)", "Status"])
        wb.save(speed_record_file_location)

# Initialize the Excel file
initialize_excel()

class EuclideanDistTracker:
    def __init__(self):
        self.center_points = {}
        self.id_count = 0

        # Time & speed tracking
        self.entry_time = {}
        self.exit_time = {}
        self.speed = {}
        self.capf = {}
        self.exceeded = 0
        self.total_vehicles = 0

    def update(self, objects_rect):
        objects_bbs_ids = []

        for rect in objects_rect:
            x, y, w, h = rect
            cx = (x + x + w) // 2
            cy = (y + y + h) // 2

            same_object_detected = False

            for id, pt in self.center_points.items():
                dist = math.hypot(cx - pt[0], cy - pt[1])

                if dist < 70:
                    self.center_points[id] = (cx, cy)
                    objects_bbs_ids.append([x, y, w, h, id])
                    same_object_detected = True

                    # Record entry time
                    if 410 <= y <= 430 and id not in self.entry_time:
                        self.entry_time[id] = time.time()

                    # Record exit time
                    if 235 <= y <= 255 and id not in self.exit_time:
                        self.exit_time[id] = time.time()

            if not same_object_detected:
                self.center_points[self.id_count] = (cx, cy)
                objects_bbs_ids.append([x, y, w, h, self.id_count])
                self.id_count += 1

        # Clean up old points
        new_center_points = {obj_bb_id[4]: self.center_points[obj_bb_id[4]] for obj_bb_id in objects_bbs_ids}
        self.center_points = new_center_points.copy()

        return objects_bbs_ids

    def get_speed(self, id):
        """Calculate the speed of the vehicle"""
        if id in self.entry_time and id in self.exit_time:
            time_diff = self.exit_time[id] - self.entry_time[id]

            if time_diff > 0:
                distance = 214.15  # meters
                speed = (distance / time_diff) * 3.6  # Convert m/s to km/h
                self.speed[id] = int(speed)
                return int(speed)

        return 0

    def capture(self, img, x, y, h, w, id):
        """Capture and save images of overspeeding vehicles"""
        if id not in self.capf:
            self.capf[id] = True

            crop_img = img[y - 5:y + h + 5, x - 5:x + w + 5]

            img_name = f"{id}_speed_{self.speed[id]}kmh.jpg"
            img_path = os.path.join(traffic_record_folder_name, img_name)
            cv2.imwrite(img_path, crop_img)

            if self.speed[id] > limit:
                exceeded_img_path = os.path.join(traffic_record_folder_name, "exceeded", img_name)
                cv2.imwrite(exceeded_img_path, crop_img)
                self.exceeded += 1

            # Append data to Excel file
            wb = load_workbook(speed_record_file_location)
            ws = wb.active

            status = "Overspeed" if self.speed[id] > limit else "Normal"
            ws.append([id, self.speed[id], status])

            wb.save(speed_record_file_location)

            self.total_vehicles += 1

    def end(self):
        """Write summary at the end of the Excel file"""
        wb = load_workbook(speed_record_file_location)
        ws = wb.active
        ws.append(["Summary", "", ""])
        ws.append(["Total Vehicles", self.total_vehicles, ""])
        ws.append(["Overspeeding", self.exceeded, ""])
        wb.save(speed_record_file_location)
